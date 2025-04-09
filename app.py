import pandas as pd   
import re
import openpyxl

source_table = '415._Сведения_по_начислениям_и_оплатам_по_домам_06 ПЖРТ.xlsx'
data_table = '2024 ПЖРТ 11.xlsx'
month = 'июнь'

def get_415_adress(street):
    street_name = re.sub(r'УЛ\.|ПР-Т|,Г.САМАРА', '', street).strip()
    street_name = street_name.lower().replace(' ', '', -1)
    return street_name

def split_values(val):
    val = str(val)
    val = val.replace(' ', '', -1).lower()
    if '/' in val:
        res = val.split('/')
        for i in range(2):
            res[i] = res[i].replace('.', '', -1)
        return res  
    else:
        return val.lower(), None

def find_columns(sheet):
    header = 10
    for col_idx, cell in enumerate(sheet[header], start=1):
        if cell.value == f'{month} план начисление\nсумма':
            income = sheet.cell(row=header, column=col_idx).column_letter
            payment = sheet.cell(row=header, column=col_idx+1).column_letter
            return income, payment
    return None, None

data = pd.read_excel(data_table, sheet_name='баланс 24', header=9).iloc[1:127]
source = pd.read_excel(source_table, sheet_name=0, header=2)

source = source[source['Услуга'] == '2Текущий ремонт общего имущества']

source['key_address'] = source['Улица'].apply(get_415_adress)
source['key_house'] = source['Дом'].apply(lambda x: split_values(x)[0])

source = source[['Текущие начисления', 'Оплачено', 'key_address', 'key_house']]

if (source['key_address'] == 'демократическая').sum() > 1:
    democratic_streets = source[source['key_address'] == 'демократическая']
    source = source[source['key_address'] != 'демократическая']
    summed_row = democratic_streets.iloc[:, [0, 1]].sum()

    new_row = democratic_streets.iloc[0].copy()
    new_row.iloc[[0, 1]] = summed_row

    new_row_df = pd.DataFrame([new_row])
    source = pd.concat([source, new_row_df], ignore_index=True)

data[['key1_address', 'key2_address']] = data['Адрес/Улица'].apply(split_values).apply(pd.Series)
data[['key1_house', 'key2_house']] = data['№ Дома'].apply(split_values).apply(pd.Series)

data['key1'] = data['key1_address'] + ' ' + data['key1_house']
data = data.drop(columns=['key1_house', 'key1_address'])

def find_full_name(address):
    for full in source['key_address']:
        if address in full:
            return full

data['key2_address'] = data['key2_address'].apply(lambda x: find_full_name(str(x)))

data['key2'] = data['key2_address'] + ' ' + data['key2_house']
data = data.drop(columns=['key2_house', 'key2_address'])

source['key'] = source['key_address'] + ' ' + source['key_house']
source = source.drop(columns=['key_address', 'key_house'])

data['check'] = data['key1'].isin(source['key'])

data['key'] = data.apply(
    lambda row: row['key1'] if row['check'] else row['key2'],
    axis=1
)

data = data.drop(columns=['check', 'key2', 'key1'])

merged = data.merge(source, on='key')

workbook = openpyxl.load_workbook(data_table)
sheet = workbook.active

income, payment = find_columns(sheet)

for row in range(12, 138):
    sheet[f'{income}{row}'].value, sheet[f'{payment}{row}'].value = merged['Текущие начисления'].iloc[row-12], merged['Оплачено'].iloc[row-12]

#workbook.save('edited.xlsx')