import pandas as pd   
import re
import openpyxl
import os

class Data:
    def __init__(self):
        self.source = None
        self.data = None
        self.data_table = ''
        self.month = ''
    
    def set_month(self, month):
        self.month = month.lower()
    
    def read_source_table(self, path):
        self.source = pd.read_excel(path, sheet_name=0, header=2)
    
    def read_data_table(self, path):
        self.data_table = path
        self.data = pd.read_excel(path, sheet_name='баланс 24', header=9).iloc[1:127]
    
    def get_result_table(self, path):

        def get_415_adress(street):
            street_name = re.sub(r'УЛ\.|ПР-Т|,Г.САМАРА', '', street).strip()
            street_name = street_name.lower().replace(' ', '', -1)
            return street_name

        def split_values(val):
            val = str(val)
            val = val.replace(' ', '', -1).lower()
            if '/' in val:
                res = val.split('/')
                for i in range(2):
                    res[i] = res[i].replace('.', '', -1)
                return res  
            else:
                return val.lower(), None

        def find_columns(sheet):
            header = 10
            for col_idx, cell in enumerate(sheet[header], start=1):
                if cell.value == f'{self.month} план начисление\nсумма':
                    income = sheet.cell(row=header, column=col_idx).column_letter
                    payment = sheet.cell(row=header, column=col_idx+1).column_letter
                    return income, payment
            return None, None

        self.source = self.source[self.source['Услуга'] == '2Текущий ремонт общего имущества']

        self.source['key_address'] = self.source['Улица'].apply(get_415_adress)
        self.source['key_house'] = self.source['Дом'].apply(lambda x: split_values(x)[0])

        self.source = self.source[['Текущие начисления', 'Оплачено', 'key_address', 'key_house']]

        if (self.source['key_address'] == 'демократическая').sum() > 1:
            democratic_streets = self.source[self.source['key_address'] == 'демократическая']
            self.source = self.source[self.source['key_address'] != 'демократическая']
            summed_row = democratic_streets.iloc[:, [0, 1]].sum()

            new_row = democratic_streets.iloc[0].copy()
            new_row.iloc[[0, 1]] = summed_row

            new_row_df = pd.DataFrame([new_row])
            self.source = pd.concat([self.source, new_row_df], ignore_index=True)

        self.data[['key1_address', 'key2_address']] = self.data['Адрес/Улица'].apply(split_values).apply(pd.Series)
        self.data[['key1_house', 'key2_house']] = self.data['№ Дома'].apply(split_values).apply(pd.Series)

        self.data['key1'] = self.data['key1_address'] + ' ' + self.data['key1_house']
        self.data = self.data.drop(columns=['key1_house', 'key1_address'])

        def find_full_name(address):
            for full in self.source['key_address']:
                if address in full:
                    return full

        self.data['key2_address'] = self.data['key2_address'].apply(lambda x: find_full_name(str(x)))

        self.data['key2'] = self.data['key2_address'] + ' ' + self.data['key2_house']
        self.data = self.data.drop(columns=['key2_house', 'key2_address'])

        self.source['key'] = self.source['key_address'] + ' ' + self.source['key_house']
        self.source = self.source.drop(columns=['key_address', 'key_house'])

        self.data['check'] = self.data['key1'].isin(self.source['key'])

        self.data['key'] = self.data.apply(
            lambda row: row['key1'] if row['check'] else row['key2'],
            axis=1
        )

        self.data = self.data.drop(columns=['check', 'key2', 'key1'])

        merged = self.data.merge(self.source, on='key')

        workbook = openpyxl.load_workbook(self.data_table)
        sheet = workbook.active

        income, payment = find_columns(sheet)

        for row in range(12, 138):
            sheet[f'{income}{row}'].value, sheet[f'{payment}{row}'].value = merged['Текущие начисления'].iloc[row-12], merged['Оплачено'].iloc[row-12]

        file_name = os.path.basename(self.data_table)

        name, ext = os.path.splitext(file_name)

        new_file_name = f"{name} (изменено){ext}"

        new_path = os.path.join(path, new_file_name)
        workbook.save(new_path)
